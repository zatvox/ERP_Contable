from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


SOURCE_FILES = [
    Path(r"C:\Users\Usuario\Downloads\balanza_de_comprobación (5).xlsx"),
    Path(r"C:\Users\Usuario\Downloads\Diario (account.journal).csv"),
    Path(r"C:\Users\Usuario\Downloads\Cuenta (account.account).csv"),
    Path(r"C:\Users\Usuario\Downloads\Reporte de Cuentas por Pagar a la Fecha (account.payable.report.current.date.line).xlsx"),
    Path(r"C:\Users\Usuario\Downloads\Reporte de Cuentas por Cobrar a la Fecha (account.receivable.report.current.date.line).xlsx"),
    Path(r"C:\Users\Usuario\Downloads\Producto (product.template).csv"),
    Path(r"C:\Users\Usuario\Downloads\Contacto (res.partner).csv"),
]

OUTPUT_DIR = Path("markdown_exports")


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^\w\s().-]+", "", value, flags=re.UNICODE)
    value = re.sub(r"\s+", "_", value)
    return value[:150].strip("._") or "archivo"


def md_escape(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, dt.date):
        text = value.isoformat()
    elif isinstance(value, dt.time):
        text = value.isoformat()
    else:
        text = str(value)
    text = text.replace("\\", "\\\\")
    text = text.replace("|", "\\|")
    text = text.replace("\r\n", "<br>")
    text = text.replace("\n", "<br>")
    text = text.replace("\r", "<br>")
    return text


def md_table(headers: Sequence[object], rows: Iterable[Sequence[object]]) -> list[str]:
    escaped_headers = [md_escape(header) or f"Columna {index + 1}" for index, header in enumerate(headers)]
    lines = [
        "| " + " | ".join(escaped_headers) + " |",
        "| " + " | ".join("---" for _ in escaped_headers) + " |",
    ]
    for row in rows:
        padded = list(row[: len(escaped_headers)])
        if len(padded) < len(escaped_headers):
            padded.extend([""] * (len(escaped_headers) - len(padded)))
        lines.append("| " + " | ".join(md_escape(cell) for cell in padded) + " |")
    return lines


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            path.read_text(encoding=encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin-1"


def read_csv_rows(path: Path) -> tuple[str, str, list[list[str]]]:
    encoding = detect_encoding(path)
    sample = path.read_text(encoding=encoding)[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    with path.open("r", encoding=encoding, newline="") as handle:
        rows = [row for row in csv.reader(handle, dialect)]
    delimiter = "\\t" if dialect.delimiter == "\t" else dialect.delimiter
    return encoding, delimiter, rows


def csv_to_markdown(path: Path, out_path: Path) -> dict[str, object]:
    encoding, delimiter, rows = read_csv_rows(path)
    max_cols = max((len(row) for row in rows), default=0)
    normalized_rows = [row + [""] * (max_cols - len(row)) for row in rows]
    headers = normalized_rows[0] if normalized_rows else []
    body = normalized_rows[1:] if normalized_rows else []

    lines = [
        f"# {path.name}",
        "",
        "## Metadatos",
        "",
        f"- Archivo origen: `{path}`",
        f"- Tipo: CSV",
        f"- Encoding detectado: `{encoding}`",
        f"- Delimitador detectado: `{delimiter}`",
        f"- Filas de datos: {len(body)}",
        f"- Columnas: {max_cols}",
        "",
        "## Datos completos",
        "",
    ]
    if headers:
        lines.extend(md_table(headers, body))
    else:
        lines.append("_Archivo sin filas._")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"rows": len(body), "columns": max_cols, "sheets": None}


def used_sheet_rows(sheet) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        rows.append([cell.value for cell in row])

    last_row = 0
    last_col = 0
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if value not in (None, ""):
                last_row = max(last_row, row_index)
                last_col = max(last_col, col_index)

    if last_row == 0 or last_col == 0:
        return []
    return [row[:last_col] for row in rows[:last_row]]


def xlsx_to_markdown(path: Path, out_path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    lines = [
        f"# {path.name}",
        "",
        "## Metadatos",
        "",
        f"- Archivo origen: `{path}`",
        f"- Tipo: XLSX",
        f"- Hojas: {len(workbook.worksheets)}",
        "",
    ]

    total_rows = 0
    max_cols = 0
    for sheet in workbook.worksheets:
        rows = used_sheet_rows(sheet)
        total_rows += len(rows)
        max_cols = max(max_cols, max((len(row) for row in rows), default=0))
        lines.extend(
            [
                f"## Hoja: {sheet.title}",
                "",
                f"- Filas incluidas: {len(rows)}",
                f"- Columnas incluidas: {max((len(row) for row in rows), default=0)}",
                "",
            ]
        )
        if rows:
            headers = [f"Columna {index + 1}" for index in range(len(rows[0]))]
            lines.extend(md_table(headers, rows))
        else:
            lines.append("_Hoja sin datos._")
        lines.append("")

    out_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return {"rows": total_rows, "columns": max_cols, "sheets": len(workbook.worksheets)}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary_lines = [
        "# Exportacion Markdown completa",
        "",
        "Estos archivos fueron convertidos sin recortar filas ni columnas detectadas.",
        "",
        "| Archivo origen | Markdown generado | Tipo | Hojas | Filas | Columnas maximas |",
        "| --- | --- | --- | --- | ---: | ---: |",
    ]

    for source in SOURCE_FILES:
        out_path = OUTPUT_DIR / f"{slugify(source.stem)}.md"
        if source.suffix.lower() == ".csv":
            stats = csv_to_markdown(source, out_path)
            file_type = "CSV"
        elif source.suffix.lower() in {".xlsx", ".xlsm"}:
            stats = xlsx_to_markdown(source, out_path)
            file_type = "XLSX"
        else:
            raise ValueError(f"Unsupported file type: {source}")

        sheets = "" if stats["sheets"] is None else str(stats["sheets"])
        summary_lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(source.name),
                    md_escape(out_path.as_posix()),
                    file_type,
                    sheets,
                    str(stats["rows"]),
                    str(stats["columns"]),
                ]
            )
            + " |"
        )

    (OUTPUT_DIR / "INDICE.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
