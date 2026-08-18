from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


REPORTS = [
    Path(r"C:\Users\Usuario\Downloads\Reporte de Cuentas por Cobrar a la Fecha (account.receivable.report.current.date.line) (1).xlsx"),
    Path(r"C:\Users\Usuario\Downloads\Reporte de Cuentas por Pagar a la Fecha (account.payable.report.current.date.line) (1).xlsx"),
]

OUTPUT_DIR = Path("markdown_exports")


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^\w\s().-]+", "", value, flags=re.UNICODE)
    value = re.sub(r"\s+", "_", value)
    return value[:170].strip("._") or "archivo"


def md_escape(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, dt.date):
        text = value.isoformat()
    elif isinstance(value, dt.time):
        text = value.isoformat()
    else:
        text = str(value)
    text = text.replace("\\", "\\\\")
    text = text.replace("|", "\\|")
    text = text.replace("\r\n", "<br>").replace("\n", "<br>").replace("\r", "<br>")
    return text


def md_table(headers: Sequence[object], rows: Iterable[Sequence[object]]) -> list[str]:
    headers = [md_escape(header) or f"Columna {index + 1}" for index, header in enumerate(headers)]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    width = len(headers)
    for row in rows:
        values = list(row[:width])
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        lines.append("| " + " | ".join(md_escape(value) for value in values) + " |")
    return lines


def used_rows(sheet) -> list[list[object]]:
    rows: list[list[object]] = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        rows.append([cell.value for cell in row])

    last_row = 0
    last_col = 0
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if value not in (None, ""):
                last_row = max(last_row, row_index)
                last_col = max(last_col, col_index)

    if last_row == 0 or last_col == 0:
        return []
    return [row[:last_col] for row in rows[:last_row]]


def is_numericish(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_group_label(text: str) -> bool:
    return bool(re.search(r"\(\d+\)\s*$", text.strip()))


def row_kind(row: Sequence[object]) -> str:
    first = "" if not row else str(row[0] or "")
    second = row[1] if len(row) > 1 else None
    if first.startswith(" ") and is_group_label(first):
        return "account_group"
    if is_group_label(first) and (is_numericish(second) or second in (None, "")):
        return "partner_group"
    return "detail"


def clean_group_name(value: object) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s*\(\d+\)\s*$", "", text).strip()


def render_grouped_section(headers: Sequence[object], body: Sequence[Sequence[object]]) -> list[str]:
    lines = [
        "## Datos reestructurados por grupos",
        "",
        "Las filas de grupo se conservaron porque contienen subtotales y conteos. En esta seccion se muestran como contexto jerarquico:",
        "",
        "- Grupo de socio: cliente/proveedor y subtotales acumulados.",
        "- Grupo de cuenta: cuenta contable dentro del socio y subtotales de esa cuenta.",
        "- Detalle: documentos o movimientos individuales.",
        "",
    ]

    current_partner = "Sin grupo de socio"
    current_account = "Sin grupo de cuenta"
    pending_details: list[Sequence[object]] = []

    def flush_details() -> None:
        nonlocal pending_details
        if pending_details:
            lines.extend(md_table(headers, pending_details))
            lines.append("")
            pending_details = []

    for row in body:
        kind = row_kind(row)
        if kind == "partner_group":
            flush_details()
            current_partner = clean_group_name(row[0])
            current_account = "Sin grupo de cuenta"
            lines.extend([f"### Socio: {md_escape(current_partner)}", "", "**Fila subtotal del socio:**", ""])
            lines.extend(md_table(headers, [row]))
            lines.append("")
        elif kind == "account_group":
            flush_details()
            current_account = clean_group_name(row[0])
            lines.extend([f"#### Cuenta: {md_escape(current_account)}", "", "**Fila subtotal de la cuenta:**", ""])
            lines.extend(md_table(headers, [row]))
            lines.extend(["", "**Detalle:**", ""])
        else:
            if current_partner == "Sin grupo de socio" and not pending_details:
                lines.extend([f"### Socio: {md_escape(current_partner)}", ""])
            if current_account == "Sin grupo de cuenta" and not pending_details:
                lines.extend([f"#### Cuenta: {md_escape(current_account)}", "", "**Detalle:**", ""])
            pending_details.append(row)

    flush_details()
    return lines


def convert_report(path: Path) -> tuple[Path, dict[str, int]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    output = OUTPUT_DIR / f"{slugify(path.stem)}_estructurado_completo.md"
    lines = [
        f"# {path.name}",
        "",
        "## Metadatos",
        "",
        f"- Archivo origen: `{path}`",
        "- Tipo: XLSX",
        f"- Hojas: {len(workbook.worksheets)}",
        "- Conversion: Markdown estructurado completo, sin omitir filas ni columnas detectadas.",
        "",
        "## Nota sobre filas agrupadas",
        "",
        "Las filas que terminan en `(n)` son utiles y no deben eliminarse si se quiere entender el reporte. Funcionan como grupos y subtotales: primero por socio y luego por cuenta contable. Las filas indentadas debajo del socio son subgrupos de cuenta; las filas restantes son el detalle del documento o movimiento.",
        "",
    ]

    total_rows = 0
    max_cols = 0
    partner_groups = 0
    account_groups = 0
    details = 0

    for sheet in workbook.worksheets:
        rows = used_rows(sheet)
        if not rows:
            lines.extend([f"## Hoja: {sheet.title}", "", "_Hoja sin datos._", ""])
            continue
        headers = rows[0]
        body = rows[1:]
        total_rows += len(rows)
        max_cols = max(max_cols, len(headers), max((len(row) for row in body), default=0))
        sheet_partner_groups = sum(1 for row in body if row_kind(row) == "partner_group")
        sheet_account_groups = sum(1 for row in body if row_kind(row) == "account_group")
        sheet_details = len(body) - sheet_partner_groups - sheet_account_groups
        partner_groups += sheet_partner_groups
        account_groups += sheet_account_groups
        details += sheet_details

        lines.extend(
            [
                f"## Hoja: {sheet.title}",
                "",
                f"- Filas totales incluidas: {len(rows)}",
                f"- Columnas incluidas: {max_cols}",
                f"- Grupos de socio detectados: {sheet_partner_groups}",
                f"- Grupos de cuenta detectados: {sheet_account_groups}",
                f"- Filas detalle detectadas: {sheet_details}",
                "",
            ]
        )
        lines.extend(render_grouped_section(headers, body))
        lines.extend(["## Tabla original completa", ""])
        lines.extend(md_table(headers, body))
        lines.append("")

    output.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return output, {
        "total_rows": total_rows,
        "max_cols": max_cols,
        "partner_groups": partner_groups,
        "account_groups": account_groups,
        "details": details,
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    index_rows = [
        "# Indice de reportes agrupados",
        "",
        "| Archivo origen | Markdown generado | Filas | Columnas maximas | Grupos socio | Grupos cuenta | Detalles |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for report in REPORTS:
        output, stats = convert_report(report)
        index_rows.append(
            "| "
            + " | ".join(
                [
                    md_escape(report.name),
                    md_escape(output.as_posix()),
                    str(stats["total_rows"]),
                    str(stats["max_cols"]),
                    str(stats["partner_groups"]),
                    str(stats["account_groups"]),
                    str(stats["details"]),
                ]
            )
            + " |"
        )
    (OUTPUT_DIR / "INDICE_REPORTES_AGRUPADOS.md").write_text("\n".join(index_rows) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
